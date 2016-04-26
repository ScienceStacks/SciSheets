"""
Wrappers for reading and writing excel files.
Creates a file IO abstraction
"""

import cell_types
import collections
import numpy as np
import openpyxl
import os
import pandas as pd

LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
NUM_LETTERS = len(LETTERS)
READ = "read"
WRITE = "write"


################### CLASSES ######################
class APIUtilExcel(object):
  """
  Class that creates an abstraction for reading
  and writing excel files. The abstraction is
  file I/O. So the use cases are:
    READ
      excel = APIUtilExcel(filepath)
      excel.openRead(<myfilepath>)
      # Set the worksheet to use
      excel.setWorksheet(<workseet name>)  # Should exist
      array = excel.readColumn(<column identifier>)
      excel.close()
    WRITE
      excel = APIUtilExcel(filepath)
      excel.openWrite(<myfilepath>)
      # Set the worksheet to use
      excel.setWorksheet(<workseet name>)  # Creates/overwrites worksheet
      excel.writeColumns(list-of-iterables, headers)
      excel.close()

  Terms
    columnid - either column letters or a column index (0 - based)
  Notes:
    1. APIUtilExec objects can be one of read or write.
    2. To read and write the same file, create two APIUtilExec objects.
  """

  def __init__(self, filepath):
    """
    :param str filepath: path to the excel file
    """
    self._filepath = filepath
    self._filemode = None  # None, READ, WRITE
    self._workbook = None
    self._worksheet = None

  def openRead(self):
    """
    Initializes connection with excel file
    :raises IOError:
    """
    self._workbook = openpyxl.load_workbook(filename=self._filepath, 
                                            read_only=True)
    self._filemode = READ

  def openWrite(self):
    """
    Initializes connection with excel file
    :raises IOError:
    """
    self._workbook = openpyxl.Workbook(write_only=True)
    self._filemode = WRITE

  def close(self):
    if self._filemode == WRITE:
      self._workbook.save(self._filepath)
    self._filemode = None
    self._workbook = None
    self._worksheet = None

  # TODO: See if this works for 3 digit letters
  def _getColumnIndex(self, columnid):
    """
    Calculate the column index from one or more letters given.
    :param ColumnID columnid:
    :return int:
    """
    if isinstance(columnid, int):
      return columnid
    # Must be column leters
    letter_list = [x.upper() for x in list(columnid)]
    index = 0
    factor = 1
    for letter in letter_list:
      index =  factor*index + LETTERS.index(letter) + 1
      factor *= NUM_LETTERS
    return index

  def setWorksheet(self, worksheet=None):
    """
    :param str worksheet: If None, the forst worksheet is used.
    :raises: KeyError if worksheet not found
    """
    if self._filemode == WRITE:
      if worksheet is not None:
        try:
          self._workbook.remove_sheet(worksheet)
        except ValueError:
          pass  # Ignore missing sheet
      else:
        worksheet = "Sheet1"
      self._worksheet = self._workbook.create_sheet(worksheet)
    elif self._filemode == READ:
      if worksheet is None:
        sheets = self._workbook.get_sheet_names()
        worksheet = sheets[0]
      self._worksheet = self._workbook[worksheet]
    else:
      raise ValueError("Must open workbook before setting worksheet.")

  def _readColumn(self, index):
    """
    :param str worksheet:
    :param int index: 1-based index of the column
    :returns list:
    """
    result = []
    for row in self._worksheet.rows:
      for cell in row:
        if cell.column == index:
          result.append(cell.value)
    return result

  def _extractHeaderFromValues(self, values, has_header):
    """
    :param list values:
    :param bool has_header: True if there is a header for the column
    :return list, str: column_values, header_value (if True)
    """
    new_values = list(values)
    if has_header:
      header = new_values[0]
      del new_values[0]
    else:
      header = None
    return new_values, header

  def readColumn(self, columnid, has_header=False):
    """
    Reads the column in the specified worksheet. If no worksheet
    is specified, then the first one is read.
    :param ColumnID columnid:
    :param bool has_header: True if there is a header for the column
    :return np.array, str: column_values, header_value (if True)
    :raises ValueError:
    """
    index = self._getColumnIndex(columnid)
    values = self._readColumn(index)
    new_values, header = self._extractHeaderFromValues(values, has_header)
    return np.array(new_values), header

  def readDataframe(self, has_header=False):
    """
    Reads the entire worksheet, returning it as a pandas dataframe.
    :param bool has_header: True if there is a header for the column
    :return pd.dataframe df: 
    :raises ValueError:
    """
    df = pd.DataFrame()
    index = 1
    while True:
      values = self._readColumn(index)
      if len(values) == 0:
        break
      new_values, header = self._extractHeaderFromValues(values, 
                                                         has_header)
      if header is None:
        header = "var%d" % index
      index += 1
      df[header] = new_values
    return df
       
  def writeColumns(self, columns, headers=None):
    """
    Writes the values into the current worksheet
    :param list-of-Iterables columns:
    :param headers list-of-str headers: hearder values to write
    :raises: ValueError
    """
    if self._worksheet is None:
      raise ValueError("Must set worksheet before writing columns.")
    if headers is not None:
      self._worksheet.append(headers)
    for column in columns:
      values = [column[idx] for column in columns]
      self._worksheet.append(values)
       
  def writeDataframe(self, df):
    """
    Writes the values into the current worksheet
    :param pd.DataFrame df:
    :raises: ValueError
    """
    if self._worksheet is None:
      raise ValueError("Must set worksheet before writing columns.")
    self._worksheet.append(list(df.columns))
    for column in df.columns:
      self._worksheet.append(list(df[column]))
